"""
Экспорт результатов в Excel (xlsx).

Назначение:
- Экспорт таблицы результатов в формат Excel.
- Форматирование заголовков, чисел и цветов.
- Автоматический выбор имени файла.

Внешние библиотеки: openpyxl.
"""
from __future__ import annotations

import logging
import os
from datetime import datetime
from typing import List, Any

logger = logging.getLogger("ui_new.export_xlsx")

# Пытаемся импортировать openpyxl
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False
    Workbook = None
    logger.warning("openpyxl не установлен. Экспорт в xlsx недоступен.")


# =============================================================================
# КОНСТАНТЫ ФОРМАТИРОВАНИЯ
# =============================================================================

# Заголовки столбцов для экспорта
EXPORT_HEADERS = [
    "Источник",
    "Метод",
    "Размер (МБ)",
    "LSD (дБ)",
    "SNR (дБ)",
    "Спектр. сх.",
    "RMSE",
    "SI-SDR (дБ)",
    "Центроид Δ (Гц)",
    "Косин. сх.",
    "Общий балл",
    "Время (с)",
    "Путь",
]

# Заголовки столбцов данных (соответствуют полям ResultRow)
EXPORT_FIELDS = [
    "source",
    "variant",
    "size_mb",
    "lsd_db",
    "snr_db",
    "spec_conv",
    "rmse",
    "si_sdr_db",
    "spec_centroid_diff_hz",
    "spec_cosine",
    "score",
    "time_sec",
    "path",
]

# Форматы чисел для каждого столбца
EXPORT_FORMATS = {
    "size_mb": "0.000",
    "lsd_db": "0.000",
    "snr_db": "0.000",
    "spec_conv": "0.000",
    "rmse": "0.00000",
    "si_sdr_db": "0.000",
    "spec_centroid_diff_hz": "0.000",
    "spec_cosine": "0.0000",
    "score": "0.0000",
    "time_sec": "0.000",
}


# =============================================================================
# ФУНКЦИИ ЭКСПОРТА
# =============================================================================

def export_results_to_xlsx(
    results: List[Any],
    output_path: str,
    title: str = "Результаты анализа",
) -> bool:
    """Экспортировать результаты в файл Excel.

    Создаёт форматированную таблицу с результатами анализа аудио.

    Параметры:
    - results: список объектов ResultRow
    - output_path: путь к файлу xlsx
    - title: заголовок листа

    Возвращает: True при успехе, False при ошибке.
    """
    if not HAS_OPENPYXL:
        logger.error("openpyxl не установлен. Экспорт невозможен.")
        return False

    try:
        # Создаём книгу
        wb = Workbook()
        ws = wb.active
        ws.title = title[:31]  # Максимум 31 символ для имени листа

        # Стили
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        
        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin")
        )

        # Заливка для разных методов (чередование цветов)
        alt_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")

        # Записываем заголовки
        for col, header in enumerate(EXPORT_HEADERS, start=1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border

        # Записываем данные
        for row_idx, r in enumerate(results, start=2):
            # Преобразуем размер в МБ
            size_mb = (getattr(r, "size_bytes", 0) or 0) / (1024 * 1024)

            # Формируем значения
            values = {
                "source": getattr(r, "source", ""),
                "variant": getattr(r, "variant", ""),
                "size_mb": size_mb,
                "lsd_db": getattr(r, "lsd_db", float("nan")),
                "snr_db": getattr(r, "snr_db", float("nan")),
                "spec_conv": getattr(r, "spec_conv", float("nan")),
                "rmse": getattr(r, "rmse", float("nan")),
                "si_sdr_db": getattr(r, "si_sdr_db", float("nan")),
                "spec_centroid_diff_hz": getattr(r, "spec_centroid_diff_hz", float("nan")),
                "spec_cosine": getattr(r, "spec_cosine", float("nan")),
                "score": getattr(r, "score", float("nan")),
                "time_sec": getattr(r, "time_sec", float("nan")),
                "path": getattr(r, "path", ""),
            }

            for col, field in enumerate(EXPORT_FIELDS, start=1):
                val = values.get(field)
                
                # Форматируем значение
                if field == "path":
                    # Путь как гиперссылка (просто текст)
                    cell = ws.cell(row=row_idx, column=col, value=val)
                elif isinstance(val, float):
                    if val != val:  # NaN check
                        cell = ws.cell(row=row_idx, column=col, value="—")
                    else:
                        cell = ws.cell(row=row_idx, column=col, value=round(val, 6))
                else:
                    cell = ws.cell(row=row_idx, column=col, value=val)

                cell.border = thin_border
                cell.alignment = Alignment(
                    horizontal="right" if field not in ("source", "variant", "path") else "left",
                    vertical="center"
                )

            # Чередование цветов строк
            if row_idx % 2 == 0:
                for col in range(1, len(EXPORT_HEADERS) + 1):
                    ws.cell(row=row_idx, column=col).fill = alt_fill

        # Автонастройка ширины столбцов
        column_widths = {
            1: 20,   # Источник
            2: 15,   # Метод
            3: 12,   # Размер
            4: 10,   # LSD
            5: 10,   # SNR
            6: 12,   # Спектр. сх.
            7: 12,   # RMSE
            8: 12,   # SI-SDR
            9: 14,   # Центроид
            10: 12,  # Косин. сх.
            11: 12,  # Балл
            12: 10,  # Время
            13: 40,  # Путь
        }

        for col, width in column_widths.items():
            ws.column_dimensions[get_column_letter(col)].width = width

        # Закрепляем первую строку
        ws.freeze_panes = "A2"

        # Сохраняем
        wb.save(output_path)
        logger.info("export_results_to_xlsx saved to %s", output_path)
        return True

    except Exception as e:
        logger.exception("export_results_to_xlsx failed: %s", e)
        return False


def generate_export_filename(prefix: str = "audio_analysis") -> str:
    """Сгенерировать имя файла для экспорта.

    Формат: {prefix}_{YYYY-MM-DD_HH-MM-SS}.xlsx
    """
    timestamp = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
    return f"{prefix}_{timestamp}.xlsx"


def is_export_available() -> bool:
    """Проверить доступность экспорта в xlsx."""
    return HAS_OPENPYXL


# =============================================================================
# ЭКСПОРТ ИМЁН
# =============================================================================

__all__ = [
    "export_results_to_xlsx",
    "generate_export_filename",
    "is_export_available",
]
